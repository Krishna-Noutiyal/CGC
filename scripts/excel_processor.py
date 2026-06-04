from scripts.csv_processor import CSVProcessor
import glob, os
import pandas as pd
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter


@dataclass
class ExcelProcessor:

    df: pd.DataFrame
    workbook: Workbook | None = field(default=None, init=False)
    worksheet: Worksheet | None = field(default=None, init=False)

    # --- Internal helpers -------------------------------------------------

    def _create_workbook(self, file_name: str) -> None:
        """
        Create an Excel workbook at the given file path.
        """
        self.workbook = Workbook()
        # Remove the default sheet created by openpyxl; we'll add our own.
        # Use ``self.workbook["Sheet"]`` (the default-sheet name) rather than
        # ``self.workbook.active`` so Pylance knows it's a real worksheet and
        # not a possibly-None chartsheet-like handle.
        self.workbook.remove(self.workbook["Sheet"])

    def _create_worksheet(self, sheet_name: str) -> None:
        """
        Create a worksheet in the workbook with the given sheet name.
        """
        if self.workbook is None:
            raise RuntimeError("_create_workbook() must be called first")
        self.worksheet = self.workbook.create_sheet(sheet_name)

    def _ws(self) -> Worksheet:
        """
        Return the active worksheet, asserting it has been created. This
        helps static type-checkers narrow ``Worksheet | None`` to ``Worksheet``.
        """
        if self.worksheet is None:
            raise RuntimeError("Worksheet has not been created yet")
        return self.worksheet

    def _add_formats(self) -> dict:
        """
        Build a dictionary of named cell styles. Each value is a dict of
        openpyxl style components (font / fill / alignment / ...) that can
        be applied to a cell via ``_apply_style``.
        """
        center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return {
            "blank": {
                "font": Font(size=14),
                "alignment": center_wrap,
            },
            "blank": {
                "font": Font(size=14),
                "alignment": center_wrap,
            },
            "orange_h": {
                "font": Font(bold=True, size=18),
                "fill": PatternFill(
                    start_color="E97132", end_color="E97132", fill_type="solid"
                ),
                "alignment": center_wrap,
            },
            "blue_h": {
                "font": Font(size=16),
                "fill": PatternFill(
                    start_color="4D93D9", end_color="4D93D9", fill_type="solid"
                ),
                "alignment": center_wrap,
            },
            "dblue_h": {
                "font": Font(size=16),
                "fill": PatternFill(
                    start_color="83CCEB", end_color="83CCEB", fill_type="solid"
                ),
                "alignment": center_wrap,
            },
            "red_h": {
                "font": Font(bold=True, size=16),
                "fill": PatternFill(
                    start_color="FF7979", end_color="FF7979", fill_type="solid"
                ),
                "alignment": center_wrap,
            },
            "green_h": {
                "font": Font(size=16),
                "fill": PatternFill(
                    start_color="00B050", end_color="00B050", fill_type="solid"
                ),
                "alignment": center_wrap,
            },
            "grey_h": {
                "font": Font(bold=True, size=18),
                "fill": PatternFill(
                    start_color="DAE9F8", end_color="DAE9F8", fill_type="solid"
                ),
                "alignment": center_wrap,
            },
            "black_h": {
                "font": Font(bold=True, size=26),
                "alignment": center_wrap,
            },
        }

    @staticmethod
    def _apply_style(cell: Cell | MergedCell, style: dict) -> None:
        """
        Apply a style dict (with optional font / fill / alignment / border /
        number_format keys) to a cell. The parameter is typed as
        ``Cell | MergedCell`` so the helper works on regular cells and on
        cells that have become part of a merged range.
        """
        if "font" in style:
            cell.font = style["font"]
        if "fill" in style:
            cell.fill = style["fill"]
        if "alignment" in style:
            cell.alignment = style["alignment"]
        if "border" in style:
            cell.border = style["border"]
        if "number_format" in style:
            cell.number_format = style["number_format"]

    def _set_cell_dimensions(self, width: float = 34.5, height: float = 48.3) -> None:
        """
        Set the default row height and column widths for the active worksheet.
        """
        ws = self._ws()
        ws.sheet_format.defaultRowHeight = height
        for col_idx in range(1, self.df.shape[1] + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Public API --------------------------------------------------------

    def Make_Excel(self, file_path: str) -> bool:
        """
        Create an Excel file at ``file_path`` containing the Capital Gains
        dashboard and the raw DataFrame data.
        """
        try:
            self._create_workbook(file_path)

            """ ##################### Capital Gains Dashboard ##################### """
            self._create_worksheet("Capital Gains")
            self._set_cell_dimensions()
            formats = self._add_formats()
            ws = self._ws()

            # Set explicit row height for the dashboard rows (1..12) so it
            # matches the original xlsxwriter ``set_default_row(48.3)`` look.
            for row_idx in range(1, 13):
                ws.row_dimensions[row_idx].height = 48.3

            """ ##################### FORMATTING CELLS ##################### """
            # Excel cell addresses are 1-indexed. Row 1 is the topmost row.

            # Row 1: SHORT TERM banner (B1:F1)
            ws.merge_cells("A1:D1")
            ws["A1"] = "SHORT TERM"
            self._apply_style(ws["A1"], formats["orange_h"])

            # Row 2: Short Term column headers
            ws["A2"] = "Full Value of Consideration"
            self._apply_style(ws["A2"], formats["blue_h"])
            ws["B2"] = "Cost of Acquisition"
            self._apply_style(ws["B2"], formats["dblue_h"])

            ws.merge_cells("E1:F1")
            ws["E1"] = "Short Term Tax"
            self._apply_style(ws["E1"], formats["grey_h"])

            # Row 6: LONG TERM banner
            ws.merge_cells("A4:D4")
            ws["A4"] = "LONG TERM"
            self._apply_style(ws["A4"], formats["orange_h"])

            # Row 7: Long Term column headers
            ws["A5"] = "Full Value of Consideration"
            self._apply_style(ws["A5"], formats["blue_h"])
            ws["B5"] = "Cost of Acquisition"
            self._apply_style(ws["B5"], formats["dblue_h"])

            ws.merge_cells("E4:F4")
            ws["E4"] = "Long Term Tax"
            self._apply_style(ws["E4"], formats["grey_h"])

            """ ##################### CALCULATING VALUES ##################### """

            short_term = self.df[self.df["Asset Type"] == "Short term"]
            long_term = self.df[self.df["Asset Type"] == "Short term"]

            fvc_short_term = short_term[
                "Sales Consideration - Reported by Source"
            ].sum()
            fvc_long_term = long_term["Sales Consideration - Reported by Source"].sum()

            coa_short_term = short_term["Cost of Acquisition"].sum()
            coa_long_term = long_term["Cost of Acquisition"].sum()

            pnl_short = fvc_short_term - coa_short_term
            pnl_long = fvc_long_term - coa_long_term

            """ ##################### INSERTING DATA INTO THE WORKSHEET ##################### """

            # Creating Formated Cells for Short Term Profit/Loss
            ws.merge_cells("C2:D2")
            if fvc_short_term - coa_short_term < 0:
                ws["C2"] = "Short Term Loss"
                self._apply_style(ws["C2"], formats["red_h"])
            else:
                ws["C2"] = "Short Term Profit"
                self._apply_style(ws["C2"], formats["green_h"])

            # Creating Formated Cells for Long TermProfit/Loss
            ws.merge_cells("C5:D5")
            if fvc_short_term - coa_short_term < 0:
                ws["C5"] = "Long Term Loss"
                self._apply_style(ws["C5"], formats["red_h"])
            else:
                ws["C5"] = "Long Term Profit"
                self._apply_style(ws["C5"], formats["green_h"])

            """ ##################### SHORT TERM VALUES ##################### """

            # Row 3: Short Term values
            ws["A3"] = fvc_short_term
            self._apply_style(ws["A3"], formats["blank"])
            ws["B3"] = coa_short_term
            self._apply_style(ws["B3"], formats["blank"])

            # Short Term Profit/Loss
            ws.merge_cells("C3:D3")
            ws["C3"] = "=A3-B3"
            self._apply_style(ws["C3"], formats["black_h"])

            # Short Term Tax (E2:F3 merged)
            ws.merge_cells("E2:F3")
            ws["E2"] = "=IF(C3*0.2<0,0,C3*0.2)"
            self._apply_style(ws["E2"], formats["black_h"])

            """ ##################### LONG TERM VALUES ##################### """

            # Row 8: Long Term values
            ws["A6"] = fvc_long_term
            self._apply_style(ws["A6"], formats["blank"])
            ws["B6"] = coa_long_term
            self._apply_style(ws["B6"], formats["blank"])

            # Long Term Profit/Loss
            ws.merge_cells("C6:D6")
            ws["C6"] = "=A6-B6"
            self._apply_style(ws["C6"], formats["black_h"])

            # Long Term Tax (E5:F7 merged)
            ws.merge_cells("E5:F6")
            ws["E5"] = "=IF(C6<=125000,0,(C6-125000)*0.125)"
            self._apply_style(ws["E5"], formats["black_h"])

            """ ##################### GRAND TOTAL OF TAX ##################### """
            # Row 11: Grand Total Tax
            ws.merge_cells("A7:B7")
            ws["A7"] = "Total Tax"
            self._apply_style(ws["A7"], formats["grey_h"])

            ws.merge_cells("C7:F7")
            ws["C7"] = "=SUM(E2,E5)"
            self._apply_style(ws["C7"], formats["black_h"])

            """ ##################### Overall Profit & Loss ##################### """

            # Row 11: Overall Profit/Loss
            ws.merge_cells("A8:B8")
            ws["A8"] = "Overall Profit/Loss"
            self._apply_style(ws["A8"], formats["grey_h"])

            ws.merge_cells("C8:F8")
            ws["C8"] = "=C3+C6"
            self._apply_style(ws["C8"], formats["red_h"])

            """ ##################### Capital Gains Data ##################### """
            # Create a new worksheet for the raw DataFrame
            self._create_worksheet("Capital Gains Data")
            self._set_cell_dimensions(width=26)
            data_ws = self._ws()
            data_formats = self._add_formats()

            # Header row height
            data_ws.row_dimensions[1].height = 55

            # Write the header
            header_format: dict = {
                "font": Font(bold=True, size=16),
                "fill": PatternFill(
                    start_color="E97132", end_color="E97132", fill_type="solid"
                ),
                "alignment": Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                ),
            }
            for col_num, col_name in enumerate(self.df.columns, start=1):
                cell = data_ws.cell(row=1, column=col_num, value=col_name)
                self._apply_style(cell, header_format)

            # Write the data rows
            for row_offset, (_, row) in enumerate(self.df.iterrows(), start=2):
                data_ws.row_dimensions[row_offset].height = 30
                for col_num, value in enumerate(row, start=1):
                    cell = data_ws.cell(row=row_offset, column=col_num)
                    if col_num == 1:
                        cell.font = Font(size=11)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                    else:
                        self._apply_style(cell, data_formats["blank"])
                    cell.value = None if pd.isna(value) else value

            # Totals row
            total_row = len(self.df) + 2
            data_ws.row_dimensions[total_row].height = 30
            for col_num, _ in enumerate(self.df.columns, start=1):
                col_letter = get_column_letter(col_num)
                cell = data_ws.cell(
                    row=total_row,
                    column=col_num,
                    value=f"=SUM({col_letter}2:{col_letter}{len(self.df) + 1})",
                )
                self._apply_style(cell, data_formats["green_h"])

            # "Total" label in the first column
            total_label = data_ws.cell(row=total_row, column=1, value="Total")
            self._apply_style(total_label, data_formats["grey_h"])

            assert self.workbook is not None
            self.workbook.save(file_path)
            return True
        except Exception as e:
            # print(f"Error creating Excel file: {str(e)}")
            return False


if __name__ == "__main__":
    # Create an instance of CSVProcessor
    test = CSVProcessor()

    # List all CSV files in the /test folder
    test_folder = "short_sale_calculator/test"
    file_list = glob.glob(os.path.join(test_folder, "*.csv"))
    print("Files found:", file_list)

    # Example usage of combine_csvs with the found files
    # df = test.combine_csvs(file_list)

    # summary = ExcelProcessor(df=df)

    # summary.Make_Excel("short_sale_calculator/test/Capital_Gains_Summary.xlsx")
